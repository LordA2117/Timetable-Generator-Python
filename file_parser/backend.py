from openpyxl import load_workbook
from os import listdir
import pprint


def readWorkbook(path) -> dict:
    wb = load_workbook(path)
    ws = wb.active
    num = ws.max_row
    compiled = {}
    for i in range(2, num + 1):
        lis = [i.value for i in ws[i]]
        compiled[lis[0]] = lis[1:]
    return compiled


# pprint(readWorkbook(
#     r".\SchoolData\Bhaswati Chattopadhyay.xlsx")
# )


def checkClashes(path1, path2, day) -> list:
    t1 = readWorkbook(path1)
    t2 = readWorkbook(path2)
    day1 = t1[day]
    day2 = t2[day]
    clashes = []
    for i in range(len(day1)):
        if day1[i] and day2[i]:
            p1 = day1[i].split('\n')
            p2 = day2[i].split('\n')
            for j in p1:
                if j in p2:
                    clashes.append(
                        f'Period {i + 1}: Clash detected. 2 teachers assigned in the same period for 1 or more groups.')
                    break

    return clashes


# res = checkClashes(r"SchoolData\Bhaswati Chattopadhyay.xlsx",
#                    r"SchoolData\Bini P Kuriakose.xlsx", 'Monday')

# pprint(res)


def viewFreeAndBusy(folder, day, period_number, view_busy=False) -> list:
    files = listdir(folder)
    freeTeachers = []
    busyTeachers = []
    for i in files:
        period = readWorkbook(f'{folder}\{i}')[
            day][period_number - 1]
        if period == None:
            freeTeachers.append(i)
        else:
            busyTeachers.append(i)

    if view_busy == False:
        return freeTeachers
    return busyTeachers


def style_worksheet(ws, cell_range):  # Styling worksheet
    from openpyxl.styles import Border, Side, Alignment
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal='center')


# counts = number of times subject shud occur per week. Must be iterable
def generate_raw(subjects, counts):
    out = []
    for i in range(len(subjects)):
        out.extend([subjects[i]]*counts[i])
    return out


# test = generate_raw(subjects, counts)  # ! Logging
# print(test)


def generateTimetable(matrix_of_periods, number_of_periods, number_of_days, repeatCount):
    from pprint import pprint
    import random
    out = []
    for i in range(number_of_days):
        out.append([None]*number_of_periods)
    # pprint(len(out))

    if len(matrix_of_periods) != (len(out)*len(out[0])):
        raise(ValueError(
            'Number of periods specified for the week doesn\'t match number of periods in the week'))

    coordinates = []
    for i in range(len(out)):
        rowNumber = i
        index = 0
        while index < len(out[rowNumber]):
            coordinates.append((rowNumber, index))
            index += 1
    # pprint(len(coordinates))

    for subject in matrix_of_periods:
        coords = random.choice(coordinates)
        # print(coords)
        if out[coords[0]].count(subject) >= repeatCount:
            coords = random.choice(coordinates)
            # print(f'Changed coordinates to {coords}')
        out[coords[0]][coords[1]] = subject
        coordinates.remove(coords)
        # print(len(coordinates))

    # pprint(out)
    return out


# * Now generates timetables using 2 different functions
def createNewTimetable(teachers_and_periods, counts, number_of_periods, number_of_days, repeatCount, filename, foldername):
    from pprint import pprint
    import string
    import xlsxwriter
    from openpyxl import load_workbook
    import json
    import os

    colLetters = [letter for letter in string.ascii_uppercase]

    days_of_the_week = ['Monday', 'Tuesday', 'Wednesday',
                        'Thursday', 'Friday', 'Saturday', 'Sunday']

    worksheet_col1 = ['Days'] + days_of_the_week[:number_of_days]

    workbook = xlsxwriter.Workbook(rf'{foldername}/{filename}.xlsx')
    worksheet = workbook.add_worksheet('Timetable')

    for i in range(number_of_days + 1):
        worksheet.write(i, 0, worksheet_col1[i])

    for i in range(1, number_of_periods + 1):
        worksheet.write(0, i, i)
    workbook.close()

    subjects = list(teachers_and_periods.keys())
    raw = generate_raw(subjects=subjects, counts=counts)
    # print(len(raw))
    newTimetable = generateTimetable(
        raw, number_of_periods, number_of_days, repeatCount)

    # pprint(newTimetable)

    raw_table = {}

    schoolDays = days_of_the_week[:number_of_days]
    for i in range(len(newTimetable)):
        raw_table[schoolDays[i]] = newTimetable[i]
    # pprint(raw_table)

    wb = load_workbook(filename=rf'{foldername}/{filename}.xlsx')
    ws = wb.active

    keys = raw_table.keys()
    rows = list(range(2, ws.max_row+1))

    for i in range(len(keys)):
        row = [1] + list(ws[rows[i]])
        row[1:] = raw_table[schoolDays[i]]
        p = ws[rows[i]]
        for j in range(1, len(p)):
            p[j].value = row[j]

    # style_worksheet(ws, f'A1:{colLetters[ws.max_column - 1]}{ws.max_row}')

    if not os.path.exists('config.json'):
        with open('config.json', mode='w') as file:
            file.write(json.dumps(teachers_and_periods))
    else:
        with open('config.json', mode='r+') as file:
            data = json.load(file)

        data.update(teachers_and_periods)

        with open('config.json', mode='w') as file:
            file.write(json.dumps(data))
    wb.save(rf'{foldername}/{filename}.xlsx')


# teacherDict = {
#     'Maths(SU)': 'Susanna Abraham',
#     'Maths(GP)': 'Ganesaperumal B',
#     'Physics(S)': 'Susan Sobi',
#     'Chemistry(B)': 'Bini P Kuriakose',
#     'Computer(J)': 'Jones Solomon Roche',
#     'PT(M)': 'Maruthupandian',
#     'Art(S)': 'Sashi',
#     'Biology(S)': 'Swami',
#     'Music(M)': 'Manuel'
# }  #? Test data

# counts = (4, 4, 8, 8, 6, 2, 2, 5, 1)

# createNewTimetable(teacherDict, counts, 8, 5, 2, 'eggs', './')


# Folder = destination folder
def createPersonalTimetable(folder, name, timetables_folder, number_of_days, number_of_periods):
    import json
    from openpyxl import load_workbook
    import xlsxwriter
    import string
    import os

    with open('config.json', mode='r') as file:
        data = json.load(file)

    colLetters = [letter for letter in string.ascii_uppercase]

    days_of_the_week = ['Monday', 'Tuesday', 'Wednesday',
                        'Thursday', 'Friday', 'Saturday', 'Sunday']

    worksheet_col1 = ['Days'] + days_of_the_week[:number_of_days]

    workbook = xlsxwriter.Workbook(rf'{folder}/{name}.xlsx')
    worksheet = workbook.add_worksheet('Timetable')

    for i in range(number_of_days + 1):
        worksheet.write(i, 0, worksheet_col1[i])

    for i in range(1, number_of_periods + 1):
        worksheet.write(0, i, i)
    workbook.close()

    files = os.listdir(rf'{timetables_folder}')

    allPeriods = []

    for file in files:
        allPeriods.append(readWorkbook(rf'{timetables_folder}\{file}'))

    personalTimetable = readWorkbook(rf'{folder}/{name}.xlsx')

    for file in files:
        filename = file.replace('.xlsx', '')
        timetable = readWorkbook(rf'{timetables_folder}\{file}')
        for day in timetable:
            l1 = timetable[day]
            l2 = personalTimetable[day]
            for i in range(len(l1)):
                if name == data[l1[i]]:  # l1[i] = [subject of teacher]
                    if l2[i] == None:
                        l2[i] = filename
                    elif l2[i] != None:
                        l2[i] += f'\n{filename}'

    wb = load_workbook(filename=rf'{folder}\{name}.xlsx')
    ws = wb.active

    keys = list(personalTimetable.keys())
    rows = list(range(2, ws.max_row+1))

    for i in range(len(keys)):
        row = [1] + list(ws[rows[i]])
        row[1:] = personalTimetable[days_of_the_week[:number_of_days][i]]
        p = ws[rows[i]]
        for j in range(1, len(p)):
            p[j].value = row[j]

    style_worksheet(ws, f'A1:{colLetters[ws.max_column - 1]}{ws.max_row}')

    wb.save(rf'{folder}\{name}.xlsx')


# createPersonalTimetable('./', "Ganesaperumal B",
#                         r'C:\Users\abhin\OneDrive\Desktop\Timetables', 5, 8)

# TODO: Add the new timetable generator to the final gui and make changes accordingly
